import os # Módulo que fornece funções para interagir com o sistema operacional
import xlrd # Módulo para ler arquivos Excel (.xls)
import openpyxl # Módulo para ler arquivos Excel (.xlsx)
from criptografar_arquivo import criptografar_arquivo_caminho # Função que criptografa um arquivo
from varredura import varredura # Função que realiza a varredura do texto extraido do arquivo em busca de nomes, CPFs, Etnias e Religiões
from print_textbox import print_to_textbox # Função que imprime mensagens em um Textbox

def processar(arquivo, textbox):
    """
    Processa um arquivo .xls ou .xlsx, procurando por nomes e CPFs.

    Argumento:
        arquivo (str): O caminho para o arquivo .xls ou .xlsx a ser processado.
        textbox (CTkTextbox): O objeto CTkTextbox onde as mensagens devem ser exibidas.
    """

    print_to_textbox(textbox, f"Processando Excel: {os.path.basename(arquivo)}",)

    texto = extrair_texto(arquivo) # Extrai texto completo do arquivo Excel
    texto_str = ' '.join(texto) # Transforma a lista de textos completos em uma única string
    
    dados_sensiveis_encontrados = varredura(textbox, texto_str, arquivo) # Verifica se há dados sensíveis no texto extraído do arquivo Excel
    if dados_sensiveis_encontrados: # Se houver dados sensíveis
        criptografar_arquivo_caminho(arquivo) # Criptografa o arquivo
        print_to_textbox(textbox, f"\nArquivo {os.path.basename(arquivo)} foi criptografado com sucesso e salvo como {os.path.basename(arquivo+'.criptografado')}")

def extrair_texto(arquivo):
    """
    Extrai texto de um arquivo Excel (.xls ou .xlsx), ignorando células vazias.

    Argumento:
        arquivo (str): O caminho para o arquivo Excel a ser processado.

    Retorna:
        list: Lista de textos extraídos do arquivo Excel.
    """
    textos = []  # Inicializa a lista para armazenar os textos do arquivo Excel
    if arquivo.endswith('.xls'):  # Para arquivos .xls
        workbook = xlrd.open_workbook(arquivo)  # Abre o arquivo Excel usando a biblioteca xlrd
        for sheet in workbook.sheets():  # Repete para as planilhas no arquivo Excel
            for row_idx in range(1, sheet.nrows):  # Repete para as linhas na planilha (excluindo a linha do cabeçalho)
                for col_idx in range(sheet.ncols):  # Repete para as colunas na planilha
                    cell_value = sheet.cell_value(row_idx, col_idx)  # Obtém o valor da célula
                    if cell_value is not None and cell_value != '':  # Verifica se a célula não está vazia
                        textos.append(str(cell_value))  # Adiciona o texto à lista
        return textos

    elif arquivo.endswith('.xlsx'):  # Para arquivos .xlsx
        workbook = openpyxl.load_workbook(arquivo)  # Abre o arquivo Excel usando a biblioteca openpyxl
        for sheet_name in workbook.sheetnames:  # Repete para os nomes das planilhas no arquivo Excel
            worksheet = workbook[sheet_name]  # Obtém a planilha atual
            for row in worksheet.iter_rows(values_only=True):  # Repete para as linhas na planilha
                for cell_value in row:  # Repete para os valores das células na linha
                    if cell_value is not None and cell_value != '':  # Verifica se a célula não está vazia
                        textos.append(str(cell_value))  # Adiciona o texto à lista
        return textos