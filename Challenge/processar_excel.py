import os
import xlrd
import openpyxl
from buscar import encontrar_nomes, encontrar_cpf, encontrar_cnpj, encontrar_etnias
from criptografar_arquivo import criptografar_arquivo_caminho
from colorama import Fore

def processar(arquivo):
    """
    Processa um arquivo .xls ou .xlsx, procurando por nomes e CPFs.

    Argumento:
        arquivo (str): O caminho para o arquivo .xls ou .xlsx a ser processado.
    """

    print(Fore.LIGHTWHITE_EX + "Processando Excel:", os.path.basename(arquivo))
    texto = extrair_texto(arquivo) # Extrai texto completo do arquivo Excel
    texto_str = ' '.join(texto) # Transforma a lista de textos completos em uma única string
    
    encontrados_nomes = encontrar_nomes(texto_str) # Encontra nomes no texto extraido do arquivo
    encontrados_cpf = encontrar_cpf(texto_str) # Encontra CPFs no texto extraido do arquivo
    encontrados_cnpj = encontrar_cnpj(texto_str) # Encontra CNPJs no texto extraido do arquivo
    encontrados_etnias = encontrar_etnias(texto_str) # Encontra Etnias no texto extraido do arquivo

    # -------------------------------------- Imprime os resultados -------------------------------------- #
    if encontrados_nomes:
        print(Fore.RED +  f"Nomes encontrados no arquivo {os.path.basename(arquivo)}\n")
        print(Fore.RESET)
    else:
        print(Fore.BLUE + f"Não foram encontrados nomes no arquivo {os.path.basename(arquivo)}\n")
        print(Fore.RESET)

    if encontrados_cpf:
        print(Fore.RED + f"CPF encontrado no arquivo {os.path.basename(arquivo)}\n")
        print(Fore.RESET)
    else:
        print(Fore.BLUE + f"Não encontrado CPFs no arquivo {os.path.basename(arquivo)}\n")
        print(Fore.RESET)

    if encontrados_cnpj:
        print(Fore.RED + f"CNPJ encontrado no arquivo {os.path.basename(arquivo)}\n")
        print(Fore.RESET)
    else:
        print(Fore.BLUE + f"Não encontrado CNPJs no arquivo {os.path.basename(arquivo)}\n")
        print(Fore.RESET)

    if encontrados_etnias:
        print(Fore.RED + f"Etnia encontrada no arquivo {os.path.basename(arquivo)}\n")
        print(Fore.RESET)
    else:
        print(Fore.BLUE + f"Não encontrada etnia no arquivo {os.path.basename(arquivo)}\n")
        print(Fore.RESET)
    # -------------------------------------- Imprime os resultados -------------------------------------- #
        
    if encontrados_nomes or encontrados_cpf or encontrados_cnpj or encontrados_etnias:
        criptografar_arquivo_caminho(arquivo)
    else:
        print(Fore.CYAN + f"Não encontrado dados sensíveis no arquivo {os.path.basename(arquivo)}")
        print(Fore.RESET)

def extrair_texto(arquivo):
    """
    Extrai texto de um arquivo Excel (.xls ou .xlsx), ignorando células vazias.

    Argumento:
        arquivo (str): O caminho para o arquivo Excel a ser processado.

    Retorna:
        list: Lista de textos extraídos do arquivo Excel.
    """
    textos = []  # Inicializa a lista para armazenar os textos do arquivo Excel

    if arquivo.endswith('.xls'):  # Para arquivos .xls
        workbook = xlrd.open_workbook(arquivo)  # Abre o arquivo Excel usando a biblioteca xlrd
        for sheet in workbook.sheets():  # Repete para as planilhas no arquivo Excel
            for row_idx in range(1, sheet.nrows):  # Repete para as linhas na planilha (excluindo a linha do cabeçalho)
                for col_idx in range(sheet.ncols):  # Repete para as colunas na planilha
                    cell_value = sheet.cell_value(row_idx, col_idx)  # Obtém o valor da célula
                    if cell_value is not None and cell_value != '':  # Verifica se a célula não está vazia
                        textos.append(str(cell_value))  # Adiciona o texto à lista
        return textos

    elif arquivo.endswith('.xlsx'):  # Para arquivos .xlsx
        workbook = openpyxl.load_workbook(arquivo)  # Abre o arquivo Excel usando a biblioteca openpyxl
        for sheet_name in workbook.sheetnames:  # Repete para os nomes das planilhas no arquivo Excel
            worksheet = workbook[sheet_name]  # Obtém a planilha atual
            for row in worksheet.iter_rows(values_only=True):  # Repete para as linhas na planilha
                for cell_value in row:  # Repete para os valores das células na linha
                    if cell_value is not None and cell_value != '':  # Verifica se a célula não está vazia
                        textos.append(str(cell_value))  # Adiciona o texto à lista
        return textos